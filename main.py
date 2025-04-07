from tkinter.font import Font
import traceback
from fastapi import FastAPI, Response, UploadFile, File
import shutil
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import pdfplumber
import re
import pandas as pd
from pathlib import Path
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from reportlab.pdfgen import canvas
import pdfkit
from fpdf import FPDF
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape  
from datetime import datetime



app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:4200"], 
    allow_credentials=True,
    allow_methods=["*"],  
    allow_headers=["*"],  
)

# Directorios
BASE_DIR = Path(__file__).resolve().parent
PDF_SIN_ANALIZAR = BASE_DIR / "pdf_sin_analizar"
PDF_ANALIZADOS = BASE_DIR / "pdf_analizados"
EXCELS_GENERADOS = BASE_DIR / "excels_generados"
PDF_GENERADOS = BASE_DIR / "pdfs_generados"
# Crear directorios si no existen
for folder in [PDF_SIN_ANALIZAR, PDF_ANALIZADOS, EXCELS_GENERADOS, PDF_GENERADOS]:
    folder.mkdir(parents=True, exist_ok=True)


# Expresiones regulares

#NOMBRE EMPRESA
razon_social = re.compile(r"Nombre del emisor:\s*(.+)")
pattern_periodo = re.compile(r"PERIODO\s+\d{2}\s(\d{2}\s/\s\d{4})")
pattern_142 = re.compile(r"VENTAS Y/O SERV. EXENTOS O NO\s[^\d\-−]*([\-−]?[\d,.]+)")
pattern_537 = re.compile(r"TOTAL CRÉDITOS\s[^\d\-−]*([\-−]?[\d,.]+)")
pattern_538 = re.compile(r"TOTAL DÉBITOS\s[^\d\-−]*([\-−]?[\d,.]+)")

@app.post("/subir_pdf/")
async def subir_pdf(file: UploadFile = File(...)):
    file_path = PDF_SIN_ANALIZAR / file.filename
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    return {"mensaje": "Archivo subido exitosamente", "nombre": file.filename}


def limpiar_numero(valor):
    if valor == "N/A":
        return 0
    return int(valor.replace('\u2212', '-').replace(',', '').replace('.', ''))


@app.get("/procesar_pdf/{filename}")
def procesar_pdf(filename: str):
    pdf_path = PDF_SIN_ANALIZAR / filename
    if not pdf_path.exists():
        return {"error": "Archivo no encontrado"}

    datos = []
    nombre_emisor = "Sin Razon Social"
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if nombre_emisor == "Sin Razon Social":
                match_razon = razon_social.search(text)
                if match_razon:
                    nombre_emisor = match_razon.group(1).strip()

            if text:
               
                match_periodo = pattern_periodo.search(text)
                periodo_str = match_periodo.group(1) if match_periodo else "Sin Periodo"

                valores_142 = pattern_142.findall(text)
                valores_537 = pattern_537.findall(text)
                match_538 = pattern_538.search(text)
                valores_538 = [match_538.group(1)] if match_538 else []

                max_length = max(len(valores_142), len(valores_537), len(valores_538))
                valores_142 += ["N/A"] * (max_length - len(valores_142))
                valores_537 += ["N/A"] * (max_length - len(valores_537))
                valores_538 += ["N/A"] * (max_length - len(valores_538))

                for i in range(max_length):
                    try:
                        val_142 = limpiar_numero(valores_142[i])
                        val_537 = limpiar_numero(valores_537[i])
                        val_538 = limpiar_numero(valores_538[i])

                        ventas_netas = (val_538 / 0.19) + val_142
                        compras_netas = val_537 / 0.19
                        ventas_netas_m = ventas_netas / 1000
                        compras_netas_m = compras_netas / 1000
                        margen = ventas_netas - compras_netas

                        datos.append({
                            
                            "PERIODO": periodo_str,
                            "538": valores_538[i],
                            "537": valores_537[i],
                            "142": valores_142[i],
                            "Ventas Netas": ventas_netas,
                            "Compras Netas": compras_netas,
                            "Ventas Netas M$": ventas_netas_m,
                            "Compras Netas M$": compras_netas_m,
                            "Margen": margen,
                        })
                    except ValueError:
                        datos.append({
                            "PERIODO": periodo_str,
                            "538": valores_538[i],
                            "537": valores_537[i],
                            "142": valores_142[i],
                            "Ventas Netas": "Error",
                            "Compras Netas": "Error",
                            "Ventas Netas M$": "Error",
                            "Compras Netas M$": "Error",
                            "Margen": "Error",
                        })

        df = pd.DataFrame(datos)


        # Convertir la columna 'PERIODO' a formato datetime para ordenar y extraer el año
        df["Fecha_Ordenable"] = df["PERIODO"].apply(lambda x: datetime.strptime(x, "%m / %Y") if x != "Sin Periodo" else datetime(1900, 1, 1))
        df["Anio"] = df["Fecha_Ordenable"].dt.year

        df = df.sort_values(by="Fecha_Ordenable")

        # Realizar el cálculo de ventas netas acumuladas reiniciando por año
        df["Ventas Netas Acumuladas"] = 0.0
        acumulado = 0
        anio_anterior = None
        for index, row in df.iterrows():
            anio_actual = row["Anio"]
            if anio_actual != anio_anterior:
                acumulado = 0  # Reiniciar el acumulador al cambiar de año
                anio_anterior = anio_actual

            if isinstance(row["Ventas Netas"], (int, float)):
                acumulado += row["Ventas Netas"]
                df.loc[index, "Ventas Netas Acumuladas"] = acumulado
            else:
                df.loc[index, "Ventas Netas Acumuladas"] = "Error"
                acumulado = "Error" 


        df["Variación Acumulada"] = None  

        for anio in df["Anio"].unique():
            anio_anterior = anio - 1

            if anio_anterior in df["Anio"].values:
                df_anio_actual = df[df["Anio"] == anio].sort_values(by="Fecha_Ordenable")
                df_anio_anterior = df[df["Anio"] == anio_anterior].sort_values(by="Fecha_Ordenable")

                for index, row in df_anio_actual.iterrows():
                    fecha_actual = row["Fecha_Ordenable"]
                    vna_actual = row["Ventas Netas Acumuladas"]

                    df_mes_anterior = df_anio_anterior[df_anio_anterior["Fecha_Ordenable"].dt.month == fecha_actual.month]

                    if not df_mes_anterior.empty:
                        vna_anio_anterior = df_mes_anterior["Ventas Netas Acumuladas"].iloc[-1]  

                        if isinstance(vna_anio_anterior, (int, float)) and vna_anio_anterior != 0:
                            variacion = ((vna_actual - vna_anio_anterior) / vna_anio_anterior) 
                            df.loc[index, "Variación Acumulada"] = round(variacion, 4)  

        df.drop(columns=["Fecha_Ordenable", "Anio"], inplace=True)

        df.replace(["N/A", "NaN", None, pd.NA], 0, inplace=True)

        columnas_a_formatear = [
            "Ventas Netas", "Compras Netas", "Ventas Netas M$", 
            "Compras Netas M$", "Margen", "Ventas Netas Acumuladas"
        ]

        for col in columnas_a_formatear:
            if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].apply(lambda x: f"$ {int(x):,}".replace(",", ".") if pd.notna(x) else x)
                
        excel_filename = f"{nombre_emisor.replace(' ', '_').replace('/', '_')}.xlsx"
        excel_path = EXCELS_GENERADOS / excel_filename

        wb = Workbook()
        ws = wb.active
        
        # Unir las celdas de la A1 a la D1
        ws.merge_cells("A1:D1")

        ws["A1"] = f"Razón Social: {nombre_emisor}"
      
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)


        col_index = df.columns.get_loc("Variación Acumulada") + 1
        col_letter = get_column_letter(col_index)  
        percent_style = NamedStyle(name="percentage_no_decimals")
        percent_style.number_format = "0%"  
        wb.add_named_style(percent_style)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                cell.style = percent_style

        green_font = Font(color="008000") 
        red_font = Font(color="FF0000")  

        rule_green = CellIsRule(operator='greaterThan', formula=['0'], font=green_font)
        rule_red = CellIsRule(operator='lessThan', formula=['0'], font=red_font)

        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', rule_green)
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', rule_red)

        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)  # Obtener la letra de la columna

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 2  # Ajustar con un margen adicional
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(excel_path)
        wb.close()
 
        shutil.move(pdf_path, PDF_ANALIZADOS / filename)

        return {"mensaje": "Procesamiento exitoso", "archivo_excel": excel_filename}


@app.get("/descargar_excel/{filename}")
def descargar_excel(filename: str):
    excel_path = EXCELS_GENERADOS / filename
    if not excel_path.exists():
        return {"error": "Archivo no encontrado"}
    
    with open(excel_path, "rb") as f:
        excel_data = f.read()

    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/descargar_pdf/{filename}")
def descargar_pdf(filename: str):
    excel_filename = filename 
    excel_path = EXCELS_GENERADOS / excel_filename
    if not excel_path.exists():
        return Response(content='{"error": "Archivo Excel no encontrado"}', status_code=404, media_type="application/json")

    pdf_filename = excel_filename.replace(".xlsx", ".pdf")
    pdf_path =  PDF_GENERADOS / pdf_filename
    pdf_path_str = str(pdf_path) # Convertir a string para ReportLab

    try:
        nombre_empresa = "Nombre no encontrado" 
        try:
            wb = load_workbook(excel_path)
            sheet = wb.active

            cell_value = sheet['A1'].value
            if isinstance(cell_value, str) and cell_value.startswith("Razón Social: "):
                nombre_empresa = cell_value.split("Razón Social: ", 1)[1]
            elif isinstance(cell_value, str): 
                 nombre_empresa = cell_value
            wb.close()
        except Exception as e_openpyxl:
            print(f"Advertencia: No se pudo leer la Razón Social desde la celda A1: {e_openpyxl}")
            nombre_empresa = excel_filename.replace(".xlsx", "").replace("_", " ")

        df = pd.read_excel(excel_path, header=2)

        df.fillna(0, inplace=True) 

        # Función de formato de porcentaje
        def format_percentage(x):
            if isinstance(x, (int, float)):
                return f"{int(round(x * 100))}%"
            return str(x) 

        col_var_acum = 'Variación Acumulada'
        if col_var_acum in df.columns:
            col_var_acum_num = col_var_acum + "_Num"
            # Convertir a numérico, los errores se vuelven NaN, luego llenar con 0
            df[col_var_acum_num] = pd.to_numeric(df[col_var_acum], errors='coerce').fillna(0)
            # Formatear la columna original (la que se mostrará) usando la numérica
            df[col_var_acum] = df[col_var_acum_num].apply(format_percentage)
        else:
            print(f"Advertencia: La columna '{col_var_acum}' no se encontró en el DataFrame.")
            col_var_acum_num = None 

        styles = getSampleStyleSheet()
        title_style = styles['h1']
        title_style.alignment = 1 

        titulo = Paragraph(f"Razón social: {nombre_empresa}", title_style)

        espacio = Spacer(1, 20) 

        df_display = df.drop(columns=[col_var_acum_num], errors='ignore') 
        table_data = [df_display.columns.tolist()] + df_display.values.tolist()

        tabla = Table(table_data) 

        estilo_base_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),         # Fondo cabecera
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),    # Color texto cabecera
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),                # Alineación general
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),               # Alineación vertical
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),      # Fuente cabecera
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),               # Padding cabecera
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),       # Fondo datos
            ('GRID', (0, 0), (-1, -1), 1, colors.black),          # Bordes
            ('FONTSIZE', (0, 0), (-1, -1), 7),                    # Tamaño fuente general (ajusta según necesites)
        ])
        tabla.setStyle(estilo_base_tabla)

        color_styles = []
        var_col_idx = df.columns.get_loc("Variación Acumulada")  
        for row_idx in range(1, len(table_data)):  
            var_value_str = str(table_data[row_idx][var_col_idx]) 
            if var_value_str.startswith("-"):  
                color_styles.append(('TEXTCOLOR', (var_col_idx, row_idx), (var_col_idx, row_idx), colors.red))
            elif not var_value_str.startswith("-"): 
                color_styles.append(('TEXTCOLOR', (var_col_idx, row_idx), (var_col_idx, row_idx), colors.green))

        tabla.setStyle(TableStyle(color_styles))

        doc = SimpleDocTemplate(pdf_path_str, pagesize=landscape(letter))

        elementos = [titulo, espacio, tabla] 
        
        doc.build(elementos)

        return FileResponse(
            path=pdf_path_str,
            media_type="application/pdf",
            filename=pdf_filename 
        )

    except Exception as e:
        return Response(content='{"error": "Ocurrió un error interno al generar el PDF"}', status_code=500, media_type="application/json")

